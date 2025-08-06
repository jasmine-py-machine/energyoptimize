from flask import Flask, request, jsonify, render_template, send_file
import pulp
import uuid
import numpy as np
import pandas as pd
import logging
import os
import json
from openpyxl import Workbook
from openpyxl.chart import PieChart, BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from io import BytesIO
from PIL import Image as PILImage
import base64

app = Flask(__name__)

# Set up logging
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

# Ensure database directory exists
DATABASE_DIR = "results"
if not os.path.exists(DATABASE_DIR):
    os.makedirs(DATABASE_DIR)

class PPCEnergyModel:
    def __init__(self, machines, activity_centers, total_area, period_hours):
        self.machines = machines
        self.activity_centers = activity_centers
        self.total_area = total_area
        self.period_hours = period_hours
        self.wattage_per_hour = 1.0

    def calculate_beta(self, center_area):
        return center_area / self.total_area if self.total_area > 0 else 0

    def calculate_direct_energy(self, power_rate, forecast_hours):
        return forecast_hours * power_rate * self.wattage_per_hour

    def calculate_indirect_energy(self, total_energy, direct_energy, beta):
        return (total_energy - direct_energy) * beta

    def calculate_total_energy_center(self, machine_id, center_id, total_energy):
        power_rate, forecast_hours, _, _ = self.machines.get(machine_id, (0, 0, "", 0))
        direct_energy = self.calculate_direct_energy(power_rate, forecast_hours)
        center_area = self.activity_centers.get(center_id, 0)
        beta = self.calculate_beta(center_area)
        indirect_energy = self.calculate_indirect_energy(total_energy, direct_energy, beta)
        total_center_energy = direct_energy + indirect_energy
        return {
            'direct_energy': direct_energy,
            'indirect_energy': indirect_energy,
            'total_center_energy': total_center_energy,
            'beta': beta
        }

    def calculate_all_centers(self, total_energy):
        results = {}
        for center_id in self.activity_centers:
            for machine_id in self.machines:
                result = self.calculate_total_energy_center(machine_id, center_id, total_energy)
                results[f"{center_id}_{machine_id}"] = result
        return results

@app.route('/')
def home():
    return render_template('index.html')

@app.route('/documentation')
def documentation():
    return render_template('documentation.html')

@app.route('/api/job_templates')
def job_templates():
    try:
        with open(os.path.join('database', 'data.json'),"r") as f:
            data = json.load(f)
        return jsonify(data)
    except Exception as e:
        logger.error("Error loading job templates: %s", str(e))
        return jsonify({'status': 'error', 'message': 'Failed to load job templates'}), 500

@app.route('/optimize', methods=['POST'])
def optimize():
    try:
        data = request.get_json()
        machines = data.get('machines', [])
        jobs = data.get('jobs', [])
        total_space = data.get('total_space')
        activity_centers = data.get('activity_centers', {})
        period_hours = data.get('period_hours')
        total_energy = data.get('total_energy')

        if not (machines and jobs and activity_centers and total_space > 0 and period_hours > 0 and total_energy > 0):
            logger.error("Invalid input data: %s", data)
            return jsonify({'status': 'error', 'message': 'All fields required with positive values.'}), 400

        machines_dict = {m['id']: (m['power'], m['processing_time'], m['name'], m['cost']) for m in machines}
        energy_model = PPCEnergyModel(machines_dict, activity_centers, total_space, period_hours)
        energy_results = energy_model.calculate_all_centers(total_energy)

        prob = pulp.LpProblem("Energy_Optimization", pulp.LpMinimize)
        x = pulp.LpVariable.dicts("x", [(i, j) for i in range(len(machines)) for j in range(len(jobs))], cat='Binary')
        s = pulp.LpVariable.dicts("s", [j for j in range(len(jobs))], lowBound=0, cat='Continuous')
        y = pulp.LpVariable.dicts("y", [(j, k) for j in range(len(jobs)) for k in range(len(jobs)) if j != k], cat='Binary')

        prob += pulp.lpSum(
            x[(i, j)] * machines[i]['power'] * jobs[j]['processing_time'] +
            sum(energy_results.get(f"C{center_id}_{machines[i]['id']}", {'indirect_energy': 0})['indirect_energy'] * x[(i, j)]
                for center_id in activity_centers)
            for i in range(len(machines)) for j in range(len(jobs))
        )

        for j in range(len(jobs)):
            prob += pulp.lpSum(x[(i, j)] for i in range(len(machines))) == 1, f"Assign_{j}"

        M = sum(job['processing_time'] for job in jobs)
        for i in range(len(machines)):
            for j in range(len(jobs)):
                for k in range(len(jobs)):
                    if j != k:
                        prob += (
                            s[j] + jobs[j]['processing_time'] <= s[k] + M * (1 - x[(i, j)]) + M * (1 - x[(i, k)]) + M * y[(j, k)]
                        ), f"NoOverlap1_{i}_{j}_{k}"
                        prob += (
                            s[k] + jobs[k]['processing_time'] <= s[j] + M * (1 - x[(i, j)]) + M * (1 - x[(i, k)]) + M * (1 - y[(j, k)])
                        ), f"NoOverlap2_{i}_{j}_{k}"

        prob += pulp.lpSum(jobs[j]['area'] * x[(i, j)] for i in range(len(machines)) for j in range(len(jobs))) <= total_space, "Total_Space"

        for j in range(len(jobs)):
            if 'deadline' in jobs[j]:
                prob += s[j] + jobs[j]['processing_time'] <= jobs[j]['deadline'], f"Deadline_{j}"

        prob.solve()

        optimized_jobs = []
        optimization_output = []
        job_costs = []
        total_cost = 0
        center_energy_results = []

        for j in range(len(jobs)):
            for i in range(len(machines)):
                if pulp.value(x[(i, j)]) > 0.5:
                    machine = machines[i]
                    job = jobs[j]
                    start_time = pulp.value(s[j]) if pulp.value(s[j]) is not None else 0
                    direct_energy = machine['power'] * job['processing_time']
                    energy_cost = direct_energy * machine['cost']
                    total_cost += energy_cost

                    optimized_jobs.append({
                        'machine': machine,
                        'job': job,
                        'start_time': start_time
                    })
                    optimization_output.append(
                        f"Job {job['name']} assigned to {machine['name']} "
                        f"at start time {start_time:.2f}, "
                        f"energy cost: {energy_cost:.2f}, "
                        f"direct energy: {direct_energy:.2f} KWH"
                    )
                    job_costs.append({
                        'id': job['id'],
                        'job_name': job['name'],
                        'machine_name': machine['name'],
                        'job_cost': energy_cost,
                        'direct_energy': direct_energy,
                        'processing_time': job['processing_time'],
                        'area': job['area'],
                        'deadline': job['deadline']
                    })

        for center_id in activity_centers:
            for machine_id in machines_dict:
                result = energy_results.get(f"{center_id}_{machine_id}", {
                    'direct_energy': 0,
                    'indirect_energy': 0,
                    'total_center_energy': 0,
                    'beta': 0
                })
                center_energy_results.append({
                    'center_id': center_id,
                    'machine_id': machine_id,
                    'direct_energy': result['direct_energy'],
                    'indirect_energy': result['indirect_energy'],
                    'total_center_energy': result['total_center_energy'],
                    'beta': result['beta']
                })

        wb = Workbook()
        ws_machines = wb.active
        ws_machines.title = 'Machines' # type: ignore
        ws_machines.append(['ID', 'Name', 'Power (KW)', 'Processing Time (hrs)', 'Cost per KWH']) # type: ignore
        for machine in machines:
            ws_machines.append([ # type: ignore
                machine['id'],
                machine['name'],
                machine['power'],
                machine['processing_time'],
                machine['cost']
            ])

        ws_jobs = wb.create_sheet('Jobs')
        ws_jobs.append(['ID', 'Job Name', 'Machine Name', 'Job Cost', 'Direct Energy (KWH)', 'Processing Time (hrs)', 'Area (m²)', 'Deadline (hrs)'])
        for job in job_costs:
            ws_jobs.append([
                job['id'],
                job['job_name'],
                job['machine_name'],
                job['job_cost'],
                job['direct_energy'],
                job['processing_time'],
                job['area'],
                job['deadline']
            ])

        ws_centers = wb.create_sheet('Work_Stations')
        ws_centers.append(['Center ID', 'Area (m²)'])
        for cid, area in activity_centers.items():
            ws_centers.append([cid, area])

        ws_params = wb.create_sheet('Global_Parameters')
        ws_params.append(['Parameter', 'Value'])
        ws_params.append(['Total Shop Floor Area (m²)', total_space])
        ws_params.append(['Total Energy (KWH)', total_energy])
        ws_params.append(['Period Hours', period_hours])

        ws_results = wb.create_sheet('Optimization_Results')
        ws_results.append(['Result'])
        ws_results.append([f"Status: {pulp.LpStatus[prob.status]}"])
        ws_results.append([f"Total Cost: {total_cost:.2f}"])
        ws_results.append([])
        for line in optimization_output:
            ws_results.append([line])

        ws_pie = wb.create_sheet('Pie_Chart_Data')
        ws_pie.append(['Job Name', 'Job Cost'])
        for job in job_costs:
            ws_pie.append([job['job_name'], job['job_cost']])
        if len(job_costs) > 0:
            pie_chart = PieChart()
            labels = Reference(ws_pie, min_col=1, min_row=2, max_row=1 + len(job_costs))
            data = Reference(ws_pie, min_col=2, min_row=1, max_row=1 + len(job_costs))
            pie_chart.add_data(data, titles_from_data=True)
            pie_chart.set_categories(labels)
            pie_chart.title = 'Cost Distribution'
            pie_chart.dataLabels = DataLabelList()
            pie_chart.dataLabels.showVal = True
            pie_chart.dataLabels.showPercent = True
            colors = ['ff6384', '36a2eb', 'ffce56', '4bc0c0', '9966ff']
            for i, series in enumerate(pie_chart.series):
                if i < len(colors):
                    series.graphicalProperties.solidFill = colors[i]
            pie_chart.width = 12
            pie_chart.height = 10
            ws_pie.add_chart(pie_chart, 'D2')

        ws_bar = wb.create_sheet('Bar_Chart_Data')
        ws_bar.append(['Job Name', 'Energy Cost'])
        for job in job_costs:
            ws_bar.append([job['job_name'], job['job_cost']])
        if len(job_costs) > 0:
            bar_chart = BarChart()
            labels = Reference(ws_bar, min_col=1, min_row=2, max_row=1 + len(job_costs))
            data = Reference(ws_bar, min_col=2, min_row=1, max_row=1 + len(job_costs))
            bar_chart.add_data(data, titles_from_data=True)
            bar_chart.set_categories(labels)
            bar_chart.title = 'Energy Costs'
            bar_chart.x_axis.title = 'Job Name'
            bar_chart.y_axis.title = 'Energy Cost'
            bar_chart.dataLabels = DataLabelList()
            bar_chart.dataLabels.showVal = True
            bar_chart.series[0].graphicalProperties.solidFill = '36a2eb'
            bar_chart.width = 12
            bar_chart.height = 10
            ws_bar.add_chart(bar_chart, 'D2')

        ws_line = wb.create_sheet('Line_Graph_Data')
        ws_line.append(['Center ID', 'Machine ID', 'Direct Energy (KWH)', 'Indirect Energy (KWH)', 'Total Energy (KWH)', 'Beta'])
        for result in center_energy_results:
            ws_line.append([
                result['center_id'],
                result['machine_id'],
                result['direct_energy'],
                result['indirect_energy'],
                result['total_center_energy'],
                result['beta']
            ])
        if len(center_energy_results) > 0:
            line_chart = LineChart()
            labels = Reference(ws_line, min_col=1, min_row=2, max_row=1 + len(center_energy_results))
            direct_data = Reference(ws_line, min_col=3, min_row=1, max_row=1 + len(center_energy_results))
            indirect_data = Reference(ws_line, min_col=4, min_row=1, max_row=1 + len(center_energy_results))
            line_chart.add_data(direct_data, titles_from_data=True)
            line_chart.add_data(indirect_data, titles_from_data=True)
            line_chart.set_categories(labels)
            line_chart.title = 'Energy Metrics'
            line_chart.x_axis.title = 'Center'
            line_chart.y_axis.title = 'Energy (KWH)'
            line_chart.series[0].graphicalProperties.line.solidFill = 'ff6384'
            line_chart.series[1].graphicalProperties.line.solidFill = '36a2eb'
            line_chart.series[0].marker.symbol = 'circle'
            line_chart.series[1].marker.symbol = 'circle'
            line_chart.width = 12
            line_chart.height = 10
            ws_line.add_chart(line_chart, 'I2')

        for ws in wb:
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter # type: ignore
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 2
                ws.column_dimensions[column].width = adjusted_width

        excel_filename = f"optimization_results_{uuid.uuid4()}.xlsx"
        wb.save(f"{DATABASE_DIR}/{excel_filename}")

        response_data = {
            'optimization_result': '\n'.join(optimization_output),
            'jobs': job_costs,
            'total_cost': total_cost,
            'center_energy_results': center_energy_results,
            'status': 'optimal' if prob.status == pulp.LpStatusOptimal else 'infeasible',
            'excel_file': excel_filename
        }

        logger.debug("Optimization successful: %s", response_data)
        return jsonify(response_data)

    except Exception as e:
        logger.error("Optimization error: %s", str(e))
        return jsonify({'status': 'error', 'message': f'Optimization failed: {str(e)}'}), 500

@app.route('/generate_pdf', methods=['POST'])
def generate_pdf():
    try:
        data = request.get_json()
        machines = data.get('machines', [])
        jobs = data.get('jobs', [])
        centers = data.get('centers', [])
        total_space = data.get('total_space')
        total_energy = data.get('total_energy')
        period_hours = data.get('period_hours')
        optimization_result = data.get('optimization_result', '')
        center_energy_results = data.get('center_energy_results', [])
        total_cost = data.get('total_cost', 0)
        chart_images = data.get('chart_images', {})

        pdf_buffer = BytesIO()
        doc = SimpleDocTemplate(pdf_buffer, pagesize=letter)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph("Energy Optimization Report", styles['Title']))
        elements.append(Spacer(1, 12))

        elements.append(Paragraph("Machines", styles['Heading2']))
        machine_data = [['ID', 'Name', 'Power (KW)', 'Processing Time (hrs)', 'Cost per KWH']] + [
            [m['id'], m['name'], m['power'], m['processing_time'], m['cost']] for m in machines
        ]
        machine_table = Table(machine_data)
        machine_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(machine_table)
        elements.append(Spacer(1, 12))

        elements.append(Paragraph("Jobs", styles['Heading2']))
        job_data = [['ID', 'Job Name', 'Machine Name', 'Job Cost', 'Direct Energy (KWH)', 'Processing Time (hrs)', 'Area (m²)', 'Deadline (hrs)']] + [
            [j['id'], j['job_name'], j['machine_name'], f"{j['job_cost']:.2f}", f"{j['direct_energy']:.2f}", j['processing_time'], j['area'], j['deadline']] for j in jobs
        ]
        job_table = Table(job_data)
        job_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(job_table)
        elements.append(Spacer(1, 12))

        elements.append(Paragraph("Work Stations", styles['Heading2']))
        center_data = [['Center ID', 'Area (m²)']] + [[c['id'], c['area']] for c in centers]
        center_table = Table(center_data)
        center_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(center_table)
        elements.append(Spacer(1, 12))

        elements.append(Paragraph("Global Parameters", styles['Heading2']))
        param_data = [
            ['Parameter', 'Value'],
            ['Total Shop Floor Area (m²)', total_space],
            ['Total Energy (KWH)', total_energy],
            ['Period Hours', period_hours]
        ]
        param_table = Table(param_data)
        param_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(param_table)
        elements.append(Spacer(1, 12))

        elements.append(Paragraph("Optimization Results", styles['Heading2']))
        result_data = [[f"Status: {pulp.LpStatus[prob.status]}"], [f"Total Cost: {total_cost:.2f}"], ['']] + [[line] for line in optimization_result.split('\n')]
        result_table = Table(result_data)
        result_table.setStyle(TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BACKGROUND', (0, 0), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(result_table)
        elements.append(Spacer(1, 12))

        elements.append(Paragraph("Pie Chart Data", styles['Heading2']))
        pie_data = [['Job Name', 'Job Cost']] + [[j['job_name'], f"{j['job_cost']:.2f}"] for j in jobs]
        pie_table = Table(pie_data)
        pie_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(pie_table)
        elements.append(Spacer(1, 12))
        if 'pie-chart' in chart_images:
            img_data = base64.b64decode(chart_images['pie-chart'].split(',')[1])
            img = PILImage.open(BytesIO(img_data))
            img_buffer = BytesIO()
            img.save(img_buffer, format="PNG")
            elements.append(Image(img_buffer, width=300, height=200))
            elements.append(Spacer(1, 12))

        elements.append(Paragraph("Bar Chart Data", styles['Heading2']))
        bar_data = [['Job Name', 'Energy Cost']] + [[j['job_name'], f"{j['job_cost']:.2f}"] for j in jobs]
        bar_table = Table(bar_data)
        bar_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(bar_table)
        elements.append(Spacer(1, 12))
        if 'bar-chart' in chart_images:
            img_data = base64.b64decode(chart_images['bar-chart'].split(',')[1])
            img = PILImage.open(BytesIO(img_data))
            img_buffer = BytesIO()
            img.save(img_buffer, format="PNG")
            elements.append(Image(img_buffer, width=300, height=200))
            elements.append(Spacer(1, 12))

        elements.append(Paragraph("Line Graph Data", styles['Heading2']))
        line_data = [['Center ID', 'Machine ID', 'Direct Energy (KWH)', 'Indirect Energy (KWH)', 'Total Energy (KWH)', 'Beta']] + [
            [r['center_id'], r['machine_id'], f"{r['direct_energy']:.2f}", f"{r['indirect_energy']:.2f}", f"{r['total_center_energy']:.2f}", f"{r['beta']:.4f}"] for r in center_energy_results
        ]
        line_table = Table(line_data)
        line_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(line_table)
        elements.append(Spacer(1, 12))
        if 'line-graph' in chart_images:
            img_data = base64.b64decode(chart_images['line-graph'].split(',')[1])
            img = PILImage.open(BytesIO(img_data))
            img_buffer = BytesIO()
            img.save(img_buffer, format="PNG")
            elements.append(Image(img_buffer, width=300, height=200))
            elements.append(Spacer(1, 12))

        doc.build(elements)
        pdf_filename = f"database/optimization_report_{uuid.uuid4()}.pdf"
        with open(pdf_filename, 'wb') as f:
            f.write(pdf_buffer.getvalue())

        return jsonify({'status': 'success', 'pdf_file': pdf_filename})

    except Exception as e:
        logger.error("PDF generation error: %s", str(e))
        return jsonify({'status': 'error', 'message': f'PDF generation failed: {str(e)}'}), 500

@app.route('/download_file/<filename>')
def download_file(filename):
    try:
        filepath = os.path.join(DATABASE_DIR, os.path.basename(filename))
        if not os.path.exists(filepath):
            logger.error("File not found: %s", filepath)
            return jsonify({'status': 'error', 'message': 'File not found'}), 404
        download_name = 'optimization_results.xlsx' if filename.endswith('.xlsx') else 'optimization_report.pdf'
        return send_file(filepath, as_attachment=True, download_name=download_name)
    except Exception as e:
        logger.error("File download error: %s", str(e))
        return jsonify({'status': 'error', 'message': 'Failed to download file'}), 500

if __name__ == '__main__':
    app.run(debug=True, port=3000, host='0.0.0.0')